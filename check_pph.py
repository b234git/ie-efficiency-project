import openpyxl

wb = openpyxl.load_workbook('c:/Users/mphat/Desktop/work/IE-Eff/info/EFF JAN V6.xlsx', data_only=True)
ws = wb['DB']

current_section = None
section_pph_map = {}

# Start from column 1 to 50
for c in range(1, 55):
    val_row1 = ws.cell(1, c).value
    if val_row1 is not None and str(val_row1).strip() != '':
        if "Unnamed" not in str(val_row1):
            current_section = str(val_row1).strip()
    
    val_row2 = ws.cell(2, c).value
    if val_row2 is not None and str(val_row2).strip() == 'PPH' and current_section:
        section_pph_map[current_section] = c
        # 0-indexed for POI: c - 1
        print(f"Section '{current_section}' has PPH at column {c} (0-indexed: {c-1})")
