import openpyxl
from openpyxl.utils import get_column_letter

wb = openpyxl.load_workbook(r'c:\Users\mphat\Desktop\work\IE-Eff\info\EFF JAN V6.xlsx', data_only=False)
ws = wb['D']

with open('formulas2.txt', 'w') as f:
    for c in range(31, 36):
        col = get_column_letter(c)
        header = ws.cell(row=5, column=c).value
        val6 = ws.cell(row=6, column=c).value
        v_formula = val6 if val6 else "None"
        f.write(f'Col {col} ({header}): {v_formula}\n')
