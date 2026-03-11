import openpyxl
import sys

file_path = r'c:\Users\mphat\Desktop\work\IE-Eff\info\EFF FEB V6.xlsx'
out_path = 'debug_extract.txt'
print(f"Loading {file_path}...")
try:
    wb = openpyxl.load_workbook(file_path, data_only=False)
except Exception as e:
    print(f"Error loading: {e}")
    sys.exit(1)

target = '398846'

with open(out_path, 'w', encoding='utf-8') as f:
    for sheet_name in wb.sheetnames:
        sheet = wb[sheet_name]
        for row in sheet.iter_rows(min_row=1, max_row=500):
            for cell in row:
                if cell.value is not None:
                    val_str = str(cell.value)
                    if target in val_str:
                        f.write(f"--- FOUND IN SHEET {sheet_name} ROW {cell.row} COL {cell.column} ---\n")
                        # get headers if row > 4
                        headers = [str(c.value) if c.value else c.column_letter for c in sheet[4]] if cell.row > 4 else []
                        for c in row:
                            if c.value is not None:
                                col_idx = c.column - 1
                                header = headers[col_idx] if col_idx < len(headers) else c.column_letter
                                val = repr(c.value)
                                f.write(f"{c.column_letter} [{header}]: {val}\n")
                        print("Written to debug_extract.txt")
                        sys.exit(0)
                        
    f.write("Not found.\n")
