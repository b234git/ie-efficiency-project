import openpyxl

wb = openpyxl.load_workbook(r'c:\Users\mphat\Desktop\work\IE-Eff\info\EFF FEB V6.xlsx', data_only=True)
sheet = wb['DB']

lines = []
for i in range(1, 40):
    val1 = sheet.cell(row=1, column=i).value
    val2 = sheet.cell(row=2, column=i).value
    lines.append(f"{openpyxl.utils.get_column_letter(i)}: {val1} -> {val2}")

print("\n".join(lines))
