"""
Import Sheet 'DB' từ EFF Excel vào bảng dbo.master_db (SQL Server).
- UPSERT theo (REF, data_month) (update nếu tồn tại, insert nếu mới)
- #DIV/0! và cell rỗng → NULL
- Bỏ qua hàng REF rỗng hoặc Article No rỗng
"""

import openpyxl
import pyodbc
import math

EXCEL_PATH = r"c:\Users\mphat\Desktop\work\IE-Eff\info\EFF JAN V6.xlsx"
DATA_MONTH = "2026-01"
CONN_STR = (
    "DRIVER={ODBC Driver 17 for SQL Server};"
    "SERVER=localhost,1433;"
    "DATABASE=ShoeEffDB;"
    "UID=sa;"
    "PWD=1;"
    "TrustServerCertificate=yes;"
)

def cell_str(cell):
    """Đọc giá trị cell thành string, trả về None nếu rỗng."""
    val = cell.value if cell else None
    if val is None:
        return None
    s = str(val).strip()
    if s == "" or s == "None":
        return None
    # Nếu là số nguyên dạng float (e.g. 46024.0) → trả về chuỗi int
    try:
        f = float(s)
        if f == int(f):
            return str(int(f))
        return s
    except (ValueError, TypeError):
        return s

def cell_float(cell):
    """Đọc giá trị cell thành float, trả về None nếu rỗng / lỗi (#DIV/0! v.v.)."""
    val = cell.value if cell else None
    if val is None:
        return None
    if isinstance(val, str):
        s = val.strip()
        if s == "" or s.startswith("#") or s == "None":
            return None
        try:
            return float(s)
        except ValueError:
            return None
    if isinstance(val, (int, float)):
        if math.isnan(val) or math.isinf(val):
            return None
        return float(val)
    return None

def main():
    print(f"Doc Excel: {EXCEL_PATH}")
    wb = openpyxl.load_workbook(EXCEL_PATH, data_only=True)
    ws = wb["DB"]

    rows_data = []
    skipped = 0

    for row_idx in range(3, ws.max_row + 1):          # bo 2 hang header
        row = ws[row_idx]

        ref_val = cell_str(row[0])                    # col A
        article_no = cell_str(row[1])                 # col B

        # Bo qua hang khong co REF hop le hoac Article No
        if not ref_val or ref_val == "0" or not article_no:
            skipped += 1
            continue

        rows_data.append({
            "ref":                   ref_val,
            "article_no":            article_no,
            "pattern_no":            cell_str(row[2]),   # C
            "shoe_name":             cell_str(row[3]),   # D
            "os_code":               cell_str(row[4]),   # E
            # SEW
            "sew_ct":                cell_float(row[5]),  # F
            "sew_mp":                cell_float(row[6]),  # G
            "sew_quota_db":          cell_float(row[7]),  # H
            "sew_pph":               cell_float(row[8]),  # I
            # BUFFING 1ST
            "buff1st_ct":            cell_float(row[9]),  # J
            "buff1st_mp":            cell_float(row[10]), # K
            "buff1st_quota_db":      cell_float(row[11]), # L
            "buff1st_pph":           cell_float(row[12]), # M
            # BUFFING 2ND
            "buff2nd_ct":            cell_float(row[13]), # N
            "buff2nd_mp":            cell_float(row[14]), # O
            "buff2nd_quota_db":      cell_float(row[15]), # P
            "buff2nd_pph":           cell_float(row[16]), # Q
            # STOCKFIT UV
            "stockfit_uv_ct":        cell_float(row[17]), # R
            "stockfit_uv_mp":        cell_float(row[18]), # S
            "stockfit_uv_quota_db":  cell_float(row[19]), # T
            "stockfit_uv_pph":       cell_float(row[20]), # U
            # STOCKFIT 1ST
            "stockfit1st_ct":        cell_float(row[21]), # V
            "stockfit1st_mp":        cell_float(row[22]), # W
            "stockfit1st_quota_db":  cell_float(row[23]), # X
            "stockfit1st_pph":       cell_float(row[24]), # Y
            # STOCKFIT 2ND
            "stockfit2nd_ct":        cell_float(row[25]), # Z
            "stockfit2nd_mp":        cell_float(row[26]), # AA
            "stockfit2nd_quota_db":  cell_float(row[27]), # AB
            "stockfit2nd_pph":       cell_float(row[28]), # AC
            # ASSEMBLY BIG
            "assem_big_ct":          cell_float(row[29]), # AD
            "assem_big_mp":          cell_float(row[30]), # AE
            "assem_big_quota_db":    cell_float(row[31]), # AF
            "assem_big_pph":         cell_float(row[32]), # AG
            # ASSEMBLY SMALL
            "assem_small_ct":        cell_float(row[33]), # AH
            "assem_small_mp":        cell_float(row[34]), # AI
            "assem_small_quota_db":  cell_float(row[35]), # AJ
            "assem_small_pph":       cell_float(row[36]), # AK
        })

    print(f"Doc duoc {len(rows_data)} hang hop le, bo qua {skipped} hang rong/zero.")

    # Connect SQL Server
    print("Ket noi SQL Server...")
    conn = pyodbc.connect(CONN_STR, autocommit=False)
    cursor = conn.cursor()

    # MERGE SQL (UPSERT theo ref + data_month)
    MERGE_SQL = """
MERGE dbo.master_db AS target
USING (SELECT ? AS ref, ? AS data_month) AS src
    ON target.ref = src.ref AND target.data_month = src.data_month
WHEN MATCHED THEN
    UPDATE SET
        article_no           = ?,
        pattern_no           = ?,
        shoe_name            = ?,
        os_code              = ?,
        sew_ct               = ?,
        sew_mp               = ?,
        sew_quota_db         = ?,
        sew_pph              = ?,
        buff1st_ct           = ?,
        buff1st_mp           = ?,
        buff1st_quota_db     = ?,
        buff1st_pph          = ?,
        buff2nd_ct           = ?,
        buff2nd_mp           = ?,
        buff2nd_quota_db     = ?,
        buff2nd_pph          = ?,
        stockfit_uv_ct       = ?,
        stockfit_uv_mp       = ?,
        stockfit_uv_quota_db = ?,
        stockfit_uv_pph      = ?,
        stockfit1st_ct       = ?,
        stockfit1st_mp       = ?,
        stockfit1st_quota_db = ?,
        stockfit1st_pph      = ?,
        stockfit2nd_ct       = ?,
        stockfit2nd_mp       = ?,
        stockfit2nd_quota_db = ?,
        stockfit2nd_pph      = ?,
        assem_big_ct         = ?,
        assem_big_mp         = ?,
        assem_big_quota_db   = ?,
        assem_big_pph        = ?,
        assem_small_ct       = ?,
        assem_small_mp       = ?,
        assem_small_quota_db = ?,
        assem_small_pph      = ?
WHEN NOT MATCHED THEN
    INSERT (ref, data_month, article_no, pattern_no, shoe_name, os_code,
            tct,
            sew_ct, sew_mp, sew_quota_db, sew_pph,
            buff1st_ct, buff1st_mp, buff1st_quota_db, buff1st_pph,
            buff2nd_ct, buff2nd_mp, buff2nd_quota_db, buff2nd_pph,
            stockfit_uv_ct, stockfit_uv_mp, stockfit_uv_quota_db, stockfit_uv_pph,
            stockfit1st_ct, stockfit1st_mp, stockfit1st_quota_db, stockfit1st_pph,
            stockfit2nd_ct, stockfit2nd_mp, stockfit2nd_quota_db, stockfit2nd_pph,
            assem_big_ct, assem_big_mp, assem_big_quota_db, assem_big_pph,
            assem_small_ct, assem_small_mp, assem_small_quota_db, assem_small_pph)
    VALUES (?, ?, ?, ?, ?, ?,
            0,
            ?, ?, ?, ?,
            ?, ?, ?, ?,
            ?, ?, ?, ?,
            ?, ?, ?, ?,
            ?, ?, ?, ?,
            ?, ?, ?, ?,
            ?, ?, ?, ?,
            ?, ?, ?, ?);
"""

    inserted = 0
    updated = 0
    errors = 0

    for i, d in enumerate(rows_data):
        data_fields = [
            d["article_no"], d["pattern_no"], d["shoe_name"], d["os_code"],
            d["sew_ct"], d["sew_mp"], d["sew_quota_db"], d["sew_pph"],
            d["buff1st_ct"], d["buff1st_mp"], d["buff1st_quota_db"], d["buff1st_pph"],
            d["buff2nd_ct"], d["buff2nd_mp"], d["buff2nd_quota_db"], d["buff2nd_pph"],
            d["stockfit_uv_ct"], d["stockfit_uv_mp"], d["stockfit_uv_quota_db"], d["stockfit_uv_pph"],
            d["stockfit1st_ct"], d["stockfit1st_mp"], d["stockfit1st_quota_db"], d["stockfit1st_pph"],
            d["stockfit2nd_ct"], d["stockfit2nd_mp"], d["stockfit2nd_quota_db"], d["stockfit2nd_pph"],
            d["assem_big_ct"], d["assem_big_mp"], d["assem_big_quota_db"], d["assem_big_pph"],
            d["assem_small_ct"], d["assem_small_mp"], d["assem_small_quota_db"], d["assem_small_pph"],
        ]

        # USING: ref + data_month (2 params)
        # UPDATE: 36 params (data_fields)
        # INSERT VALUES: ref + data_month + data_fields (38 params)
        params = (
            [d["ref"], DATA_MONTH]  # USING clause: ref + data_month
            + data_fields           # WHEN MATCHED UPDATE: 36
            + [d["ref"], DATA_MONTH]  # INSERT VALUES: ref + data_month
            + data_fields           # INSERT VALUES: data_fields (36)
        )

        try:
            cursor.execute(MERGE_SQL, params)
        except Exception as e:
            print(f"  ERROR row {i+1} REF={d['ref']}: {e}")
            errors += 1
            conn.rollback()
            conn.autocommit = True
            conn.autocommit = False

    try:
        conn.commit()
        print("Commit OK.")
    except Exception as e:
        print(f"Commit error: {e}")
        conn.rollback()

    cursor.execute("SELECT COUNT(*) FROM dbo.master_db")
    total = cursor.fetchone()[0]

    conn.close()
    print(f"\n=== RESULT ===")
    print(f"Rows processed : {len(rows_data)}")
    print(f"Errors         : {errors}")
    print(f"Total in DB    : {total}")

if __name__ == "__main__":
    main()
