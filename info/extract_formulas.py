import openpyxl
import traceback

try:
    path = r'c:\Users\mphat\Desktop\work\IE-Eff\info\EFF JAN V6.xlsx'
    
    wb = openpyxl.load_workbook(path, data_only=False)
    sheet_d = wb['D']
    
    print("--- FIRST few rows formulas ---")
    formulas_found = 0
    # Let's check row 3 (which likely has data/formulas)
    for col_idx in range(1, 100):
        cell_val = sheet_d.cell(row=3, column=col_idx).value
        # if string and starts with =
        if isinstance(cell_val, str) and cell_val.startswith('='):
            header = sheet_d.cell(row=1, column=col_idx).value or sheet_d.cell(row=2, column=col_idx).value or str(col_idx)
            print(f"Col {col_idx} ({header}): {cell_val}")
            formulas_found += 1
            if formulas_found > 30:
                break
        
except Exception as e:
    print(f"Error: {e}")
    traceback.print_exc()
