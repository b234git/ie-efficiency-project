import openpyxl
import json
import traceback

try:
    path = r'c:\Users\mphat\Desktop\work\IE-Eff\info\EFF JAN V6.xlsx'
    
    # 1. Read DB sheet (data only)
    wb_data = openpyxl.load_workbook(path, data_only=True)
    sheet_db = wb_data['DB']
    
    headers = [cell.value for cell in sheet_db[1]]
    
    # Take first 5 rows of data for inspection
    data = []
    for row in sheet_db.iter_rows(min_row=2, max_row=6, values_only=True):
        row_dict = dict(zip(headers, row))
        data.append(row_dict)
        
    print("--- DB SHEET HEADERS ---")
    print(headers)
    print("\n--- DB SHEET DATA (first 5 rows) ---")
    print(json.dumps(data, default=str, indent=2))
    
    # 2. Extract formulas from Sheet 'D' 
    # The user says "truy xuất tất cả công thức trong file này" 
    wb_formulas = openpyxl.load_workbook(path, data_only=False)
    sheet_d = wb_formulas['D']
    
    formulas = []
    # Print headers of Sheet D
    d_headers = [cell.value for cell in sheet_d[1]]
    print("\n--- SHEET D HEADERS ---")
    print((d_headers))
    
    for row in sheet_d.iter_rows(min_row=2, max_row=3):
        for cell in row:
            if cell.data_type == 'f':
                formulas.append(f"Cell {cell.coordinate}: {cell.value}")
                
    print("\n--- FORMULAS IN SHEET D (first rows) ---")
    for f in formulas:
        print(f)
        
except Exception as e:
    print(f"Error: {e}")
    traceback.print_exc()
