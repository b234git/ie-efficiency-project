import openpyxl

wb = openpyxl.load_workbook(r'c:\Users\mphat\Desktop\work\IE-Eff\info\EFF JAN V6.xlsx', data_only=True)
if 'D' in wb.sheetnames:
    ws = wb['D']
    with open(r'c:\Users\mphat\Desktop\work\IE-Eff\d_output.txt', 'w', encoding='utf-8') as f:
        for row in range(1, 8):
            row_vals = [str(ws.cell(row=row, column=c).value) for c in range(1, 26)]
            f.write(f'Row {row}: {row_vals}\n')
else:
    print('Sheet D not found')
