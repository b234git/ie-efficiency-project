import openpyxl
import sys

try:
    wb=openpyxl.load_workbook(r'c:\Users\mphat\Desktop\work\IE-Eff\info\EFF JAN V6.xlsx', data_only=False)
    ws=wb['D']
    headers = [ws.cell(row=5, column=c).value for c in range(29, 36)]
    formulas = [ws.cell(row=6, column=c).value for c in range(29, 36)]
    
    with open('d_formulas.txt', 'w', encoding='utf-8') as f:
        f.write(f'Headers: {headers}\n')
        f.write(f'Formulas: {formulas}\n')
        
    print("DONE writing to d_formulas.txt")
except Exception as e:
    with open('d_formulas.txt', 'w', encoding='utf-8') as f:
        f.write(f'Error: {e}\n')
    print("ERROR writing to d_formulas.txt")
