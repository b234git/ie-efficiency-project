import openpyxl

wb = openpyxl.load_workbook(r'c:\Users\mphat\Desktop\work\IE-Eff\info\EFF FEB V6.xlsx', data_only=True)
sheet = wb['DB']

headers = []
for i in range(1, 100):
    val = sheet.cell(row=2, column=i).value
    if val:
        headers.append(f"{openpyxl.utils.get_column_letter(i)}: {val}")

print("\n".join(headers))
