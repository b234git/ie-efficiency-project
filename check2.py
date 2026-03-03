import openpyxl
wb = openpyxl.load_workbook('c:/Users/mphat/Desktop/work/IE-Eff/info/EFF JAN V6.xlsx')
ws = wb['%']
for r in range(1, 5):
    for c in range(1, 40):
        if str(ws.cell(r, c).value).strip() == 'EFF':
            print(f'Found EFF at r={r} c={c}')
            print(f'Formula 1: {ws.cell(r+1, c).value}')
            print(f'Formula 2: {ws.cell(r+2, c).value}')
