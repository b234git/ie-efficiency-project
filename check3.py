import openpyxl

wb = openpyxl.load_workbook('c:/Users/mphat/Desktop/work/IE-Eff/info/EFF JAN V6.xlsx')
ws = wb['D']

# AE is likely index 31 (A=1, Z=26, AA=27, AB=28, AC=29, AD=30, AE=31)
print("Header of AE:", ws.cell(2, 31).value)
print("Formula at row 3 AE:", ws.cell(3, 31).value)
print("Formula at row 4 AE:", ws.cell(4, 31).value)
