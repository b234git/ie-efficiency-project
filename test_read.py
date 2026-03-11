import openpyxl
import sys

file_path = r'c:\Users\mphat\Desktop\work\IE-Eff\info\EFF FEB V6.xlsx'
print(f"Loading {file_path}...")
try:
    wb = openpyxl.load_workbook(file_path, data_only=False)
except Exception as e:
    print(f"Error loading: {e}")
    sys.exit(1)

print("Sheets:", wb.sheetnames)
target = '398846'

for sheet_name in wb.sheetnames:
    print(f"Searching sheet {sheet_name}...")
    sheet = wb[sheet_name]
    for row in sheet.iter_rows(min_row=1, max_row=500):
        for cell in row:
            if cell.value is not None:
                val_str = str(cell.value)
                if target in val_str:
                    print(f"--- FOUND IN SHEET {sheet_name} ROW {cell.row} COL {cell.column} ---")
                    for c in row:
                        if c.value is not None:
                            print(f"{c.column_letter}: {c.value}")
                    sys.exit(0)
print("Not found.")
